from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "prototype" / "dashboard-data.js"

GENERIC_SECTION_TITLES = {
    "\u65f6\u95f4\u5e8f\u5217",
    "\u65f6\u70b9",
    "\u65f6\u95f4\u533a\u95f4",
    "\u65f6\u95f4\u8303\u56f4",
}

BUSINESS_CHANGE_PAGE = "\u4e1a\u52a1\u53d8\u52a8\u5206\u6790"
BUSINESS_CHANGE_BLOCKS = {
    "\u5b58\u91cf\u4e1a\u52a1",
    "\u65b0\u53d1\u751f\u4e1a\u52a1",
    "\u5230\u671f\u4e1a\u52a1\uff08\u542b\u9759\u6001\u672a\u6765\u5230\u671f\uff09",
}

MERGED_AREA_SCOPE = "\u5408\u5e76\u533a\u57df"

LEGACY_TEXT_REPLACEMENTS = (
    (
        "\u65f6\u95f4\u5e8f\u5217\uff08\u5386\u53f2\u6708\u9891+\u5f53\u6708\u7684\u65e5\u9891\uff09",
        "\u6708\u9891 / \u65e5\u9891",
    ),
    (
        "\u65f6\u95f4\u5e8f\u5217\uff08\u5386\u53f2\u6708\u9891+\u5f53\u6708\u65e5\u9891\uff09",
        "\u6708\u9891 / \u65e5\u9891",
    ),
    (
        "\u65f6\u95f4\u5e8f\u5217\uff08\u5386\u53f2\u6708\u9891+\u672c\u6708\u9010\u65e5\uff09",
        "\u6708\u9891 / \u65e5\u9891",
    ),
    (
        "\u65f6\u95f4\u5e8f\u5217\uff08\u6708\u9891+\u5f53\u6708\u7684\u65e5\u9891\uff09",
        "\u6708\u9891 / \u65e5\u9891",
    ),
    (
        "\u65f6\u95f4\u5e8f\u5217\uff08\u6708\u9891+\u5f53\u6708\u65e5\u9891\uff09",
        "\u6708\u9891 / \u65e5\u9891",
    ),
    (
        "\u5386\u53f2\u533a\u95f4\u6309\u6708\u5c55\u793a\uff0c\u5f53\u6708\u6309\u65e5\u5c55\u793a",
        "\u652f\u6301\u5728\u56fe\u5361\u5185\u5207\u6362\u6708\u9891 / \u65e5\u9891\u5c55\u793a",
    ),
    (
        "\u5386\u53f2\u533a\u95f4\u6309\u6708\u5c55\u793a\uff0c\u672c\u6708\u6309\u65e5\u5c55\u793a",
        "\u652f\u6301\u5728\u56fe\u5361\u5185\u5207\u6362\u6708\u9891 / \u65e5\u9891\u5c55\u793a",
    ),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export dashboard workbook rows into dashboard-data.js.")
    parser.add_argument(
        "--workbook",
        type=Path,
        help="Path to the source workbook. Defaults to the first .xlsx in project root.",
    )
    parser.add_argument("--output", type=Path, default=OUTPUT, help="Output file path.")
    return parser.parse_args()


def detect_workbook(cli_path: Path | None = None) -> Path:
    if cli_path:
        workbook_path = cli_path if cli_path.is_absolute() else ROOT / cli_path
        if workbook_path.exists():
            return workbook_path
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    candidates = [
        path
        for path in ROOT.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            "Workbook not found under project root. Put the source .xlsx in the repo root or pass --workbook."
        )

    preferred = [path for path in candidates if any(ord(char) > 127 for char in path.name)]
    return sorted(preferred or candidates, key=lambda path: path.name)[0]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_legacy_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    normalized = text
    for old, new in LEGACY_TEXT_REPLACEMENTS:
        normalized = normalized.replace(old, new)
    return normalized


def split_filters(text: Any) -> list[str]:
    normalized = clean_text(text)
    if not normalized:
        return []

    normalized = normalized.replace("\uff1b", ";").replace("\uff0c", ",")
    parts = []
    for raw_part in re.split(r"[;]+", normalized):
        part = raw_part.strip().strip(",")
        if part:
            parts.append(part)
    return list(OrderedDict.fromkeys(parts))


def normalize_structural_label(value: Any) -> str:
    return normalize_legacy_text(value).replace("\u3000", " ").strip()


def is_generic_section_title(value: Any) -> bool:
    normalized = normalize_structural_label(value)
    return normalized in GENERIC_SECTION_TITLES


def normalize_view_scope(value: Any) -> str:
    return normalize_structural_label(value)


def normalize_area_name(value: Any) -> str:
    normalized = normalize_structural_label(value)
    if not normalized or is_generic_section_title(normalized):
        return ""
    return normalized


def load_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[-1]]
    rows = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue

        (
            seq,
            page_name,
            block_name,
            area_name,
            view_scope,
            widget_title,
            shared_filters,
            display_desc,
            component_type,
            grain,
            default_filters,
            frontend_params,
            axis_desc,
            metric_desc,
            legend_desc,
            response_fields,
            linkage_rule,
            dev_note,
            origin_pos,
        ) = row[:19]

        rows.append(
            {
                "seq": int(seq),
                "page_name": normalize_structural_label(page_name),
                "block_name": normalize_structural_label(block_name),
                "area_name": normalize_area_name(area_name),
                "view_scope": normalize_view_scope(view_scope),
                "widget_title": normalize_legacy_text(widget_title),
                "shared_filters": split_filters(shared_filters),
                "display_desc": normalize_legacy_text(display_desc),
                "component_type": normalize_legacy_text(component_type),
                "grain": normalize_legacy_text(grain),
                "default_filters": normalize_legacy_text(default_filters),
                "frontend_params": normalize_legacy_text(frontend_params),
                "axis_desc": normalize_legacy_text(axis_desc),
                "metric_desc": normalize_legacy_text(metric_desc),
                "legend_desc": normalize_legacy_text(legend_desc),
                "response_fields": normalize_legacy_text(response_fields),
                "linkage_rule": normalize_legacy_text(linkage_rule),
                "dev_note": normalize_legacy_text(dev_note),
                "origin_pos": normalize_legacy_text(origin_pos),
            }
        )

    return rows


def resolve_area_descriptor(row: dict[str, Any]) -> tuple[str, str, str]:
    page_name = row["page_name"]
    block_name = row["block_name"]
    area_name = row["area_name"]
    view_scope = row["view_scope"]

    if page_name == BUSINESS_CHANGE_PAGE and block_name in BUSINESS_CHANGE_BLOCKS:
        return block_name, MERGED_AREA_SCOPE, block_name

    if not area_name:
        if view_scope and not is_generic_section_title(view_scope):
            area_base = block_name or view_scope
            return area_base, view_scope, f"{area_base}__{view_scope}"
        return block_name, MERGED_AREA_SCOPE, block_name

    if not view_scope:
        return area_name, MERGED_AREA_SCOPE, area_name

    return area_name, view_scope, f"{area_name}__{view_scope}"


def merge_list(values: list[str], incoming: list[str]) -> list[str]:
    ordered = OrderedDict.fromkeys(values)
    for item in incoming:
        ordered[item] = None
    return list(ordered.keys())


def build_widget(row: dict[str, Any]) -> dict[str, Any]:
    widget: dict[str, Any] = {
        "seq": row["seq"],
        "title": row["widget_title"],
        "componentType": row["component_type"],
        "displayDescription": row["display_desc"],
        "grain": row["grain"],
        "defaultFilters": row["default_filters"],
        "frontendParams": row["frontend_params"],
        "axisDescription": row["axis_desc"],
        "metricDescription": row["metric_desc"],
        "legendDescription": row["legend_desc"],
        "responseFields": row["response_fields"],
        "linkageRule": row["linkage_rule"],
        "devNote": row["dev_note"],
        "originPosition": row["origin_pos"],
    }
    return {key: value for key, value in widget.items() if value not in (None, "")}


def build_payload(workbook_path: Path) -> dict[str, Any]:
    rows = load_rows(workbook_path)
    pages: OrderedDict[str, dict[str, Any]] = OrderedDict()

    for row in rows:
        page = pages.setdefault(
            row["page_name"],
            {"id": f"page-{len(pages) + 1}", "name": row["page_name"], "blocks": OrderedDict()},
        )
        block = page["blocks"].setdefault(
            row["block_name"],
            {"id": f"block-{len(page['blocks']) + 1}", "name": row["block_name"], "areas": OrderedDict()},
        )

        area_name, view_scope, area_key = resolve_area_descriptor(row)
        area = block["areas"].setdefault(
            area_key,
            {
                "id": f"area-{row['seq']}",
                "name": area_name,
                "viewScope": view_scope,
                "sharedFilters": [],
                "widgets": [],
            },
        )
        area["sharedFilters"] = merge_list(area["sharedFilters"], row["shared_filters"])
        area["widgets"].append(build_widget(row))

    page_list = []
    total_widgets = 0
    for page in pages.values():
        block_list = []
        for block in page["blocks"].values():
            area_list = []
            for area in block["areas"].values():
                area["widgets"].sort(key=lambda widget: widget["seq"])
                if not area["sharedFilters"]:
                    area.pop("sharedFilters", None)
                area_list.append(area)
            area_list.sort(key=lambda area: area["widgets"][0]["seq"])
            block_list.append(
                {
                    "id": block["id"],
                    "name": block["name"],
                    "areas": area_list,
                    "widgetCount": sum(len(area["widgets"]) for area in area_list),
                }
            )
        widget_count = sum(block["widgetCount"] for block in block_list)
        total_widgets += widget_count
        page_list.append(
            {
                "id": page["id"],
                "name": page["name"],
                "blocks": block_list,
                "blockCount": len(block_list),
                "areaCount": sum(len(block["areas"]) for block in block_list),
                "widgetCount": widget_count,
            }
        )

    return {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "workbook": workbook_path.name,
        "pageCount": len(page_list),
        "widgetCount": total_widgets,
        "pages": page_list,
    }


def main() -> None:
    args = parse_args()
    workbook_path = detect_workbook(args.workbook)
    payload = build_payload(workbook_path)

    output_path = args.output if args.output.is_absolute() else ROOT / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    js = "window.dashboardData = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    output_path.write_text(js, encoding="utf-8")
    print(f"Exported to {output_path}")


if __name__ == "__main__":
    main()
